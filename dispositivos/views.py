from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from .models import Device, Measurement, Alert, Zone, Category
from django.db.models import Count
from datetime import timedelta
from django.utils import timezone
from django.shortcuts import render
from usuarios.decorators import admin_required, manager_required, editor_required, reader_required
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
import openpyxl
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.auth.models import User
from usuarios.models import UserProfile
from .forms import DeviceForm, MeasurementForm, CategoryForm, ZoneForm, SensorForm, AlertForm
from dispositivos.models import Zone, Device, Category
from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count
from dispositivos.models import Zone, Device, Category, Alert, Measurement



@login_required
def dashboard(request):
    # Intentar obtener la organización del usuario
    organization = getattr(getattr(request.user, "userprofile", None), "organization", None)

    if not organization:
        zones = Zone.objects.all()
        devices = Device.objects.all()
        categories = Category.objects.all()
        alerts = Alert.objects.all()
        measurements = Measurement.objects.all()
    else:
        zones = Zone.objects.filter(organization=organization)
        devices = Device.objects.filter(organization=organization)
        categories = Category.objects.filter(organization=organization)
        alerts = Alert.objects.filter(device__organization=organization)
        measurements = Measurement.objects.filter(device__organization=organization)

    # 🔹 Zonas con cantidad de dispositivos
    zones_with_devices = zones.annotate(device_count=Count('devices')).order_by('-device_count')

    # 🔹 Alertas de la última semana
    week_ago = timezone.now() - timedelta(days=7)
    alerts_week = alerts.filter(created_at__gte=week_ago)
    alert_counts = {
        'grave': alerts_week.filter(level='GRAVE').count(),
        'alta': alerts_week.filter(level='ALTA').count(),
        'media': alerts_week.filter(level='MEDIA').count(),
    }

    # 🔹 Últimas 10 mediciones (con fecha, dispositivo, valor)
    latest_measurements = measurements.select_related('device').order_by('-date')[:10]

    context = {
        'zones_with_devices': zones_with_devices,
        'devices': devices,
        'categories': categories,
        'alert_counts': alert_counts,
        'latest_measurements': latest_measurements,
    }
    return render(request, 'dashboard.html', context)

@login_required
def device_list(request):
    try:
        organization = request.user.userprofile.organization
        if not hasattr(organization, 'pk'):
            organization = None
    except:
        organization = None

    category_filter = request.GET.get('category')
    search_query = request.GET.get('search')
    sort_by = request.GET.get('sort', 'name')

    if organization:
        devices = Device.objects.filter(organization=organization)
        if category_filter:
            devices = devices.filter(category__id=category_filter)
        if search_query:
            devices = devices.filter(
                Q(name__icontains=search_query) |
                Q(reference__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(zone__name__icontains=search_query)
            )
        devices = devices.order_by(sort_by)
        categories = Category.objects.filter(organization=organization)
    else:
        devices = Device.objects.none()
        categories = Category.objects.none()

    # Pagination
    paginator = Paginator(devices, 10)  # Show 10 devices per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    contexto = {
        'page_obj': page_obj,
        'categories': categories,
        'selected_category': category_filter,
        'search_query': search_query,
        'sort_by': sort_by,
    }
    return render(request, "device_list.html", contexto)

@login_required
def panel(request):
    return render(request, "panel.html", {})  # Crea luego el template panel.html

@login_required
def device_detail(request, pk):
    try:
        organization = request.user.userprofile.organization
        if not hasattr(organization, 'pk'):
            organization = None
    except:
        organization = None

    if organization:
        device = get_object_or_404(Device, id=pk, organization=organization)
        measurements = Measurement.objects.filter(device=device).order_by('-date')
        alerts = Alert.objects.filter(device=device).order_by('-created_at')
    else:
        device = None
        measurements = []
        alerts = []

    contexto = {
        'device': device,
        'measurements': measurements,
        'alerts': alerts,
    }
    return render(request, "device_detail.html", contexto)

from django.shortcuts import redirect

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    else:
        return redirect('login')

@login_required
def measurement_list(request):
    try:
        organization = request.user.userprofile.organization
        if not hasattr(organization, 'pk'):
            organization = None
    except:
        organization = None

    search_query = request.GET.get('search')
    sort_by = request.GET.get('sort', '-date')

    if organization:
        measurements = Measurement.objects.filter(device__organization=organization).select_related('device').order_by(sort_by)
        if search_query:
            measurements = measurements.filter(
                Q(device__name__icontains=search_query) |
                Q(value__icontains=search_query) |
                Q(unit__icontains=search_query)
            )
    else:
        measurements = Measurement.objects.none()

    # Pagination
    paginator = Paginator(measurements, 20)  # Show 20 measurements per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    contexto = {
        'page_obj': page_obj,
        'search_query': search_query,
        'sort_by': sort_by,
    }
    return render(request, "measurement_list.html", contexto)

def inicio(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    else:
        return redirect('login')

# Role-based access decorators
def admin_required(view_func):
    return user_passes_test(lambda u: u.groups.filter(name='Admin').exists())(view_func)

def manager_required(view_func):
    return user_passes_test(lambda u: u.groups.filter(name='Admin').exists() or u.groups.filter(name='Manager').exists())(view_func)


@admin_required
def admin_dashboard(request):
    # Admin specific dashboard
    total_users = User.objects.count()
    total_devices = Device.objects.count()
    total_measurements = Measurement.objects.count()
    total_alerts = Alert.objects.count()

    contexto = {
        'total_users': total_users,
        'total_devices': total_devices,
        'total_measurements': total_measurements,
        'total_alerts': total_alerts,
    }
    return render(request, "admin_dashboard.html", contexto)

@login_required
def alert_list(request):
    organization = getattr(getattr(request.user, "userprofile", None), "organization", None)

    if not organization:
        alerts = Alert.objects.all().order_by('-created_at')
    else:
        alerts = Alert.objects.filter(device__organization=organization).order_by('-created_at')

    return render(request, 'alerts.html', {'alerts': alerts})




@login_required
@manager_required
def device_create(request):
    if request.method == 'POST':
        form = DeviceForm(request.POST)
        if form.is_valid():
            device = form.save(commit=False)
            device.organization = request.user.userprofile.organization
            device.save()
            messages.success(request, 'Device created successfully.')
            return redirect('device_list')
    else:
        form = DeviceForm()
    return render(request, 'device_form.html', {'form': form, 'title': 'Create Device'})

@login_required
@manager_required
def device_update(request, pk):
    device = get_object_or_404(Device, pk=pk, organization=request.user.userprofile.organization)
    if request.method == 'POST':
        form = DeviceForm(request.POST, instance=device)
        if form.is_valid():
            form.save()
            messages.success(request, 'Device updated successfully.')
            return redirect('device_detail', pk=device.pk)
    else:
        form = DeviceForm(instance=device)
    return render(request, 'device_form.html', {'form': form, 'title': 'Update Device'})

@login_required
@manager_required
def device_delete(request, pk):
    device = get_object_or_404(Device, pk=pk, organization=request.user.userprofile.organization)
    if request.method == 'POST':
        device.delete()
        messages.success(request, 'Device deleted successfully.')
        return redirect('device_list')
    return render(request, 'device_confirm_delete.html', {'device': device})

@login_required
@manager_required
def measurement_create(request):
    if request.method == 'POST':
        form = MeasurementForm(request.POST)
        if form.is_valid():
            measurement = form.save(commit=False)
            measurement.organization = request.user.userprofile.organization
            measurement.save()
            messages.success(request, 'Measurement created successfully.')
            return redirect('measurement_list')
    else:
        form = MeasurementForm()
    return render(request, 'measurement_form.html', {'form': form, 'title': 'Create Measurement'})

@login_required
@manager_required
def measurement_update(request, pk):
    measurement = get_object_or_404(Measurement, pk=pk, device__organization=request.user.userprofile.organization)
    if request.method == 'POST':
        form = MeasurementForm(request.POST, instance=measurement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Measurement updated successfully.')
            return redirect('measurement_list')
    else:
        form = MeasurementForm(instance=measurement)
    return render(request, 'measurement_form.html', {'form': form, 'title': 'Update Measurement'})

@login_required
@manager_required
def measurement_delete(request, pk):
    measurement = get_object_or_404(Measurement, pk=pk, device__organization=request.user.userprofile.organization)
    if request.method == 'POST':
        measurement.delete()
        messages.success(request, 'Measurement deleted successfully.')
        return redirect('measurement_list')
    return render(request, 'measurement_confirm_delete.html', {'measurement': measurement})

@login_required
def export_measurements_excel(request):
    measurements = Measurement.objects.filter(device__organization=request.user.userprofile.organization).select_related('device')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurements"

    # Headers
    headers = ['Device', 'Value', 'Unit', 'Date']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Data
    for row_num, measurement in enumerate(measurements, 2):
        ws.cell(row=row_num, column=1, value=measurement.device.name)
        ws.cell(row=row_num, column=2, value=float(measurement.value))
        ws.cell(row=row_num, column=3, value=measurement.unit)
        ws.cell(row=row_num, column=4, value=measurement.date.strftime('%Y-%m-%d %H:%M:%S'))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=measurements.xlsx'
    wb.save(response)
    return response

@login_required
def export_devices_excel(request):
    devices = Device.objects.filter(organization=request.user.userprofile.organization).select_related('category', 'zone')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devices"

    # Headers
    headers = ['Name', 'Category', 'Zone', 'Reference', 'Status']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Data
    for row_num, device in enumerate(devices, 2):
        ws.cell(row=row_num, column=1, value=device.name)
        ws.cell(row=row_num, column=2, value=device.category.name)
        ws.cell(row=row_num, column=3, value=device.zone.name)
        ws.cell(row=row_num, column=4, value=device.reference)
        ws.cell(row=row_num, column=5, value=device.status)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=devices.xlsx'
    wb.save(response)
    return response
